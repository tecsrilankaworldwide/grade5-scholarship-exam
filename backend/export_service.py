"""
Export service for generating Excel and PDF reports
"""
import io
from datetime import datetime
from typing import List, Dict, Any
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not installed - Excel export disabled")

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("reportlab not installed - PDF export disabled")

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to Excel and PDF"""
    
    @staticmethod
    def export_exam_results_to_excel(
        exam_title: str,
        grade: str,
        results: List[Dict[str, Any]],
        skill_areas: List[str]
    ) -> bytes:
        """
        Export exam results to Excel file
        
        Args:
            exam_title: Title of the exam
            grade: Grade level
            results: List of student results
            skill_areas: List of skill area names
            
        Returns:
            bytes: Excel file content
        """
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Exam Results"
        
        # Header styling
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"{exam_title} - {grade}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Export date
        ws.merge_cells('A2:E2')
        date_cell = ws['A2']
        date_cell.value = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        date_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Student Name', 'Email', 'Score', 'Percentage', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Data rows
        for row_idx, result in enumerate(results, 5):
            ws.cell(row=row_idx, column=1, value=result.get('student_name', ''))
            ws.cell(row=row_idx, column=2, value=result.get('email', ''))
            ws.cell(row=row_idx, column=3, value=result.get('score', 0))
            ws.cell(row=row_idx, column=4, value=f"{result.get('percentage', 0)}%")
            ws.cell(row=row_idx, column=5, value=result.get('status', ''))
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.read()
    
    @staticmethod
    def export_student_report_to_pdf(
        student_name: str,
        grade: str,
        month: str,
        exams_data: List[Dict[str, Any]],
        skill_breakdown: Dict[str, float],
        overall_stats: Dict[str, Any]
    ) -> bytes:
        """
        Export student monthly report to PDF
        
        Args:
            student_name: Student's name
            grade: Grade level
            month: Month/year
            exams_data: List of exam results
            skill_breakdown: Performance by skill area
            overall_stats: Overall statistics
            
        Returns:
            bytes: PDF file content
        """
        if not PDF_AVAILABLE:
            raise RuntimeError("reportlab not installed")
        
        # Create PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#F97316'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1F2937'),
            spaceBefore=12,
            spaceAfter=6
        )
        
        # Title
        story.append(Paragraph("Education Reforms Bureau", title_style))
        story.append(Paragraph("Monthly Progress Report", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        # Student info
        info_data = [
            ['Student:', student_name],
            ['Grade:', grade],
            ['Month:', month],
            ['Generated:', datetime.now().strftime('%Y-%m-%d')]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Overall statistics
        story.append(Paragraph("Overall Performance", heading_style))
        stats_data = [
            ['Metric', 'Value'],
            ['Exams Taken', str(overall_stats.get('exams_taken', 0))],
            ['Average Score', f"{overall_stats.get('average_score', 0)}%"],
            ['Highest Score', f"{overall_stats.get('highest_score', 0)}%"],
            ['Lowest Score', f"{overall_stats.get('lowest_score', 0)}%"],
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F97316')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Skill breakdown
        story.append(Paragraph("Skill Area Performance", heading_style))
        skill_data = [['Skill Area', 'Performance (%)']]
        for skill, score in skill_breakdown.items():
            skill_data.append([skill, f"{score}%"])
        
        skill_table = Table(skill_data, colWidths=[4*inch, 2*inch])
        skill_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(skill_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Exam history
        story.append(Paragraph("Exam History", heading_style))
        exam_data = [['Exam Title', 'Date', 'Score', '%']]
        for exam in exams_data:
            exam_data.append([
                exam.get('title', '')[:40],
                exam.get('date', ''),
                f"{exam.get('score', 0)}/{exam.get('total', 60)}",
                f"{exam.get('percentage', 0)}%"
            ])
        
        exam_table = Table(exam_data, colWidths=[3*inch, 1.5*inch, 1*inch, 1*inch])
        exam_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(exam_table)
        
        # Build PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return pdf_buffer.read()


# Quick availability check
def check_export_availability():
    """Check which export formats are available"""
    return {
        "excel": EXCEL_AVAILABLE,
        "pdf": PDF_AVAILABLE
    }
